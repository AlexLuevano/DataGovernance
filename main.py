from openpyxl.reader.excel import load_workbook
import pandas as pd
import numpy as np
from tkinter import *
from tkinter import filedialog
import openpyxl
import pyodbc
from typing import Optional,List
import sys
import os

def browse_files():
       global filename
       filename = filedialog.askopenfilename(initialdir = "/Downloads", title = "Selecciona un archivo de MDM:", filetypes = (("all files","*.*"),(".csv Files","*.csv*")))
       if filename:
              l1 = Label(window, text = "File path: " + filename).pack()
       else:
              print('No seleccionaste ningún archivo.')
       window.destroy()

def export_file():
    global export_file_path
    export_file_path = filedialog.asksaveasfilename(defaultextension = '.xlsm',initialdir = "/Desktop", title = "Guardar archivo como:", filetypes = (("Excel Files","*.xlsm*"),(".csv Files","*.csv*"),("all files","*.*")))
    with pd.ExcelWriter(export_file_path,engine='openpyxl') as writer:
        mdm.to_excel(writer, sheet_name='Entities', index=False)
        query.to_excel(writer, sheet_name='Query',index=False)
    workbook = writer.book
    #base = os.path.splitext(export_file_path)[0]
    #new_name = os.rename(export_file_path,base+'.xlsm')
    #workbook.filename = new_name
    #workbook.add_vba_project('./vbaProject.bin')
    writer.save
    wb = load_workbook(export_file_path)
    ent_sheet = wb['Entities']
    ent_sheet.insert_rows(1)
    #writer = pd.ExcelWriter()
    #mdm.to_excel(writer, sheet_name = 'Entities', index = False, header=True)
    #query.to_excel(writer, sheet_name = 'Query', index = False, header = True)
    print('File was succesfully filled and saved on the chosen path')
    window2.destroy()

def db_connect() -> pd.DataFrame:
    print('Connecting to Aurora...')
    try:
        conn = pyodbc.connect(
            Driver='{SQL Server}',
            Server='pr-aurora1-rs01',
            Database='db_product_integrity',
            Trusted_Connection='yes')
        print('Connected successfully to Aurora')
    except:
        raise Exception('Connection to Aurora failed. Ending execution.')
        sys.exit(1)
    sql = ("""
        SELECT DISTINCT 
        CAST(A.sku AS VARCHAR) + '|' +  CAST(V.AP_REF AS VARCHAR) + '|' +  A.vendor_id AS [key2]    
        , CAST(A.sku AS VARCHAR) + '|' +  CAST(V.AP_REF AS VARCHAR)  [key]
        , A.sku
        , V.AP_REF
        , A.vendor_id
        , A.core_return_vendor
        , A.recall_vendor
        , A.warranty_vendor
        , REPLACE(A.warehouses, ' ', '') warehouses
        , ISNULL(B.rank, 999) VENDOR_RANK
        , CASE WHEN A.vendor_id != A.core_return_vendor OR A.vendor_id != A.recall_vendor OR A.vendor_id != A.warranty_vendor THEN 'Yes' ELSE 'No' END [Diff Contracts]
        , IT.part_number
        , IT.line_code
        , IT.alt_part_number
        , IT.alt_line_code
        , IT.team_cm_name
        , IT.major_dept
        , IT.minor_dept
        , IT.znet_manager
        , IT.team_cm_name
        , '' as [IM.prod_image]
        , '' as [IM.prod_attribute]
        , VD.max_per_car
        , VD.item_size as [Item Size]
        , VD.unit_of_measure [UOM for EDI Ordering]
        , VD.max_per_car [Max per Car]
        , VD.store_credit_flag [Store Credit Flag]
        , VD.oil_gallons [Oil Gal]
        , VD.[employee_discount]  [Employee Discount]
        , VD.warranty
        , VD.[warranry_months] [Warranty Months]
        , VD.[quantity_force_flag] [Quantity Force Flag]
        , COALESCE(VD.[custom_id], 'N') [Custom Id]
        , VD.[quantity_per_case] [Quantity per Case]
        , VD.schedule_b_code [Sched B Code]
        , 'No' [Michigan Flag]		
        , COALESCE(COO.coo, '') [Country of Origin]
        , CASE WHEN VD.order_pack = 0 THEN 1 ELSE VD.order_pack END [store order pack]
        , LEN(A.warehouses) L	
        , VD.order_pack	
        , VD.supply_chain_analyst
        FROM db_product_integrity.prod.tb_sku_warehouses A WITH(NOLOCK)
        LEFT JOIN dz_production.dbo.dz_item_file IT WITH(NOLOCK)
        ON A.sku = IT.item
        LEFT JOIN db_product_integrity.prod.tb_az_vendors V WITH(NOLOCK)
        ON A.vendor_id = V.VENDOR 
        LEFT JOIN db_product_integrity.prod.tb_item_vendor_rank B WITH(NOLOCK)
        ON A.sku = B.item
        AND A.vendor_id = b.vendor_id
        LEFT JOIN [db_product_integrity].[prod].[tb_sku_validation_data_upload] VD
        ON A.sku = VD.sku
        AND A.vendor_id = VD.[vendor_id]
        LEFT JOIN [db_product_integrity].[prod].[tb_sku_vendor_coo] COO
        ON A.sku = COO.sku
        AND A.vendor_id = COO.pov
        WHERE  A.SKU IN({})
        ORDER BY A.sku, VENDOR_RANK, L DESC, [Country of Origin] DESC""").format(skus)
    query = pd.read_sql(sql,conn)
    print('Query run successful. Closing connection...')
    conn.close()
    print('Closed connection to Aurora successfully')
    return query

window = Tk()
window.title("Data Governance")
window.geometry("200x100")
label_file_explorer = Label(window, text = "Data Governance", width = 100, height = 4).pack()
button_explore = Button(window, text = "Buscar archivo", command = browse_files).pack()
window.mainloop()

print('Reading MDM file...')
try:
        #wb = openpyxl.load_workbook(f'{filename}',keep_vba=True)
        #ws = wb['Entities']
        mdm = pd.read_excel(f'{filename}',sheet_name='Entities', header = 1, index_col = False,na_filter= False)
        skus_list = list(mdm['Stock Keeping Unit'])
        skus_list_string = str(skus_list)
        skus = skus_list_string[1:-1]
        print('MDM file read successfully...')
except:
        raise Exception('Error at reading MDM file. Please check file and try again. Finishing execution.')
        sys.exit(1)

mdm['KEY'] = '=CONCATENATE(INDEX(A:AAB, ROW(), MATCH("Stock Keeping Unit",$2:$2, 0)), "|",MID(INDEX(A:AAB, ROW(), MATCH("Vendor Ownership ID",$2:$2, 0)), 1, FIND("-", INDEX(A:AAB, ROW(), MATCH("Vendor Ownership ID",$2:$2, 0)))-1))'
mdm['KEY2'] = '=CONCATENATE(INDEX(A:AAB,ROW(),MATCH("Stock Keeping Unit",$2:$2,0)),"|",MID(INDEX(A:AAB,ROW(),MATCH("Vendor Ownership ID",$2:$2,0)),1,FIND("-",INDEX(A:AAB,ROW(),MATCH("Vendor Ownership ID",$2:$2,0)))-1),"|",INDEX(A:AAB,ROW(),MATCH("Vendor DCs.POV ID",$2:$2,0)))'
mdm['Ap Ref'] = '=IF(EXACT(SUBSTITUTE(INDEX(A:ZZ, ROW(), MATCH("Vendor Ownership ID",$2:$2, 0)),"-USA", ""), INDEX(A:ZZ, ROW(), MATCH("APREF",$2:$2, 0))), "TRUE", "Fix AP Ref")'
mdm['HVRPIV'] = '=VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY",$2:$2, 0)), Query!B:J, 4, 0)'
mdm['ProperPOV'] = '=IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Vendor DCs.POV ID",$2:$2, 0)), INDIRECT(ADDRESS(ROW(),COLUMN()-1))), "TRUE", "Replace POV")'
mdm['VendorRank'] = '=VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:K, 10, 0)'
mdm['DVC?'] = '=VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:K, 11, 0)'
mdm['POVs Sync'] = '=IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Vendor DCs.POV ID",$2:$2, 0)),INDEX(A:ZZ, ROW(), MATCH("Core Vendor Id",$2:$2, 0))), IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Vendor DCs.POV ID",$2:$2, 0)),INDEX(A:ZZ, ROW(), MATCH("Recall Vendor Id",$2:$2, 0))), IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Vendor DCs.POV ID",$2:$2, 0)),INDEX(A:ZZ, ROW(), MATCH("Warranty Vendor Id",$2:$2, 0))),"Ok", "Check Warranty VId"), "Check Recall VId"), "Check Core VId")'
mdm['Check Core Vendor Id'] = '=IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Core Vendor Id",$2:$2, 0)), VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:F, 6, 0)), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:F, 6, 0))'
mdm['Check Recall Vendor Id'] = '=IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Recall Vendor Id",$2:$2, 0)), VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:G, 7, 0)), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:G, 7, 0))'
mdm['Check Warranty Vendor Id'] = '=IF(EXACT(INDEX(A:ZU, ROW(), MATCH("Warranty Vendor Id",$2:$2, 0)), VLOOKUP(INDEX(A:ZU, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:H, 8, 0)), "TRUE", VLOOKUP(INDEX(A:ZU, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:H, 8, 0))'
mdm['DCs'] = '=IF(EXACT(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:K, 9, 0),INDEX(A:ZZ, ROW(), MATCH("Vendor DCs.DC\'s",$2:$2, 0))), "Ok", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:K, 9, 0))'
mdm['Major'] = '=IF(EXACT(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:Q, 17, 0), INDEX(A:ZY, ROW(), MATCH("Major Department",$2:$2, 0))), "TRUE",VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:Q, 17, 0))'
mdm['Minor'] = '=IF(EXACT(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 18, 0), INDEX(A:ZZ, ROW(), MATCH("Minor Department",$2:$2, 0))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 18, 0))'
mdm['AZ Part Number'] = '=IF(EXACT(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 12, 0), INDEX(A:ZZ, ROW(), MATCH("AutoZone Part Number",$2:$2, 0))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 12, 0))'
mdm['Line Code'] = '=IF(EXACT(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 13, 0), INDEX(A:ZZ, ROW(), MATCH("Line Code",$2:$2, 0))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 13, 0))'
mdm['Alt Part Number'] = '=IF(EXACT(TRIM(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 14, 0)), TRIM( INDEX(A:ZZ, ROW(), MATCH("Alternate Part Number",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 14, 0))'
mdm['Alt Line Code'] = '=IF(EXACT(TRIM(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 15, 0)), TRIM( INDEX(A:ZZ, ROW(), MATCH("Alternate Line Code",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:R, 15, 0))'
mdm['Max per Car'] = '=IF(EXACT(TRIM(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:W, 23, 0)), TRIM( INDEX(A:ZZ, ROW(), MATCH("Quantity per Application",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:W, 23, 0))'
mdm['Check Store Order Pack'] = '=IF(NUMBERVALUE(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AL, 38, 0)) > 1, "Check with CM-"&VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AL, 38, 0), IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Store Order Pack",$2:$2, 0)), VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AL, 38, 0)), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AL, 38, 0)))'
mdm['Check Store Credit Flag'] = '=IF(EXACT(IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,27,0)="N","No",IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,27,0)="Y","Yes","EMPTY")),INDEX(A:ZZ,ROW(),MATCH("Store Credit Flag",$2:$2,0))),"TRUE",IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,27,0)="N","No",IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,27,0)="Y","Yes","EMPTY")))'
mdm['Check Oil gal'] = '=IF(INDEX(A:ZZ, ROW(), MATCH("Oil gal",$2:$2, 0))="", "Fill out with 0.000",   IF(EXACT(NUMBERVALUE(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)),Query!A:AB, 28, 0)),NUMBERVALUE(INDEX(A:ZZ, ROW(), MATCH("Oil gal",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AB, 28, 0)))'
mdm['Check Emp Disc'] = '=IF(EXACT(IF(VLOOKUP(INDEX(A:AAC,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,29,0)="N","No",IF(VLOOKUP(INDEX(A:AAC,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,29,0)="Y","Yes","EMPTY")),INDEX(A:AAC,ROW(),MATCH("Employee Discount",$2:$2,0))),"TRUE",VLOOKUP(INDEX(A:AAC,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AC,29,0))'
mdm['Check Warranty'] = '=IF(EXACT(IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AD,30,0)="N","No",IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AD,30,0)="Y","Yes","EMPTY")),INDEX(A:ZZ,ROW(),MATCH("Warranty",$2:$2,0))),"TRUE",VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AD,30,0))'
mdm['Check Warranty Months'] = '=IF(INDEX(A:ZZ, ROW(), MATCH("Warranty Months",$2:$2, 0))="", "Fill out with 0",   IF(EXACT(NUMBERVALUE(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)),Query!A:AE, 31, 0)),NUMBERVALUE(INDEX(A:ZZ, ROW(), MATCH("Warranty Months",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AE, 31, 0)))'
mdm['Check Quantity Force Flag'] = '=IF(EXACT(IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AF,32,0)="N","No",IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AF,32,0)="Y","Yes","EMPTY")),INDEX(A:ZZ,ROW(),MATCH("Quantity Force Flag",$2:$2,0))),"TRUE",VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AF,32,0))'
mdm['Check Sched B'] = '=IF(ISBLANK(INDEX(A:ZZ,ROW(),MATCH("Harmonized Tariff Code (Schedule B)",$2:$2,0))), VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AI,35,0),IF(NUMBERVALUE( INDEX(A:ZZ,ROW(),MATCH("Harmonized Tariff Code (Schedule B)",$2:$2,0)))=0,"Clean Schedule B Code",IF(INDEX(A:ZZ,ROW(),MATCH("Harmonized Tariff Code (Schedule B)",$2:$2,0))<>"","TRUE",IF(EXACT(NUMBERVALUE(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AI,35,0)), NUMBERVALUE(INDEX(A:ZZ,ROW(),MATCH("Harmonized Tariff Code (Schedule B)",$2:$2,0)))),"TRUE",VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AI,35,0)))))'
mdm['Check Michigan Flag'] = '=IF(ISBLANK(INDEX(A:ZZ,ROW(),MATCH("Michigan Flag",$2:$2,0))),"Set Michigan Flag to No", IF(  OR(INDEX(A:ZZ,ROW(),MATCH("Michigan Flag",$2:$2,0))="No",INDEX(A:ZZ,ROW(),MATCH("Michigan Flag",$2:$2,0))="Yes"),"TRUE","Set Michigan Flag to No"))'
mdm['Check Country of Origin'] = '=IF(ISBLANK(INDEX(A:ZZ,ROW(),MATCH("Country of Origin (Primary)",$2:$2,0))), IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AK,37,0)=0,"Need COO","TRUE"),"TRUE")'
mdm['Check Custom Id'] = '=IF(EXACT(IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AO,33,0)="C", "C", "N"), INDEX(A:ZZ,ROW(),MATCH("Custom Id",$2:$2,0))), "TRUE", IF(VLOOKUP(INDEX(A:ZZ,ROW(),MATCH("KEY2",$2:$2,0)),Query!A:AO,33,0)="C", "C", "N"))'
mdm['Check Marketing Flag'] = '=IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("Marketing Flag",$2:$2, 0)), "Yes"), IF(EXACT(INDEX(A:ZZ, ROW(), MATCH("VendorRank",$2:$2, 0)),1), "TRUE", "Set MKT Flag to No"), IF(NUMBERVALUE(INDEX(A:ZZ, ROW(), MATCH("VendorRank",$2:$2, 0)))>1,"TRUE","Set MKT Flag toYes")'
mdm['Check Supply Chain Analyst'] = '=IF(EXACT(TRIM(VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AO, 41, 0)), TRIM( INDEX(A:ZZ, ROW(), MATCH("Supply Chain Analyst",$2:$2, 0)))), "TRUE", VLOOKUP(INDEX(A:ZZ, ROW(), MATCH("KEY2",$2:$2, 0)), Query!A:AO, 41, 0))'
mdm['Remarks'] = ''
print('Analysis columns added...')
query = db_connect()

# File export
window2 = Tk()
window2.title("Data Governance")
window2.geometry("200x100")
label_file_explorer = Label(window2, text = "Data Governance", width = 100, height = 4).pack()
button_explore = Button(window2, text = "Exportar archivo", command = export_file).pack()
window2.mainloop()

